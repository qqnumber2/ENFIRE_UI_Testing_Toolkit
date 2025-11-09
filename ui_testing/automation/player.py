import json
import logging
import time
import threading
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Tuple, Optional, Set, Sequence
import re
from datetime import datetime
import pyautogui
from PIL import Image

# NEW: guarded pywinauto import (so EXE still runs even if not installed)

try:
    from pywinauto import Desktop  # UIA selector
    from pywinauto.findwindows import ElementNotFoundError, WindowNotFoundError  # type: ignore
except Exception:
    Desktop = None  # type: ignore
    ElementNotFoundError = WindowNotFoundError = None  # type: ignore
# EXE-safe imports

try:
    from ui_testing.automation.action import Action  # noqa: F401
    from ui_testing.automation.util import dotted_code_from_test_name, ensure_png_name
except Exception:
    try:
        from .action import Action  # noqa: F401
        from .util import dotted_code_from_test_name, ensure_png_name
    except Exception:
        from action import Action  # type: ignore  # noqa: F401
        from util import dotted_code_from_test_name, ensure_png_name  # type: ignore
try:
    from ui_testing.automation.explorer import ExplorerController
except Exception:
    try:
        from .explorer import ExplorerController
    except Exception:
        ExplorerController = None  # type: ignore
try:
    from ui_testing.automation.driver import (
        PywinautoUnavailableError,
        AutomationSession,
        WindowSpec,
        DEFAULT_WINDOW_SPEC,
    )
    from ui_testing.automation.semantic import SemanticContext, get_semantic_context
except Exception:
    AutomationSession = None  # type: ignore
    PywinautoUnavailableError = RuntimeError  # type: ignore
    SemanticContext = None  # type: ignore

    class _FallbackWindowSpec:  # pragma: no cover - fallback when pywinauto missing
        def __init__(self, title_regex: Optional[str] = None, class_name: Optional[str] = None) -> None:
            self.title_regex = title_regex
            self.class_name = class_name

    WindowSpec = _FallbackWindowSpec  # type: ignore
    DEFAULT_WINDOW_SPEC = _FallbackWindowSpec()  # type: ignore

    def get_semantic_context(*args, **kwargs):  # type: ignore
        raise PywinautoUnavailableError("Semantic context unavailable")
try:
    from ui_testing.automation.flake_tracker import FlakeTracker
except Exception:
    FlakeTracker = None  # type: ignore
from ui_testing.automation.player_components.screenshots import ScreenshotComparator, ScreenshotResult
from ui_testing.automation.player_components.metrics import PlaybackMetrics
try:
    from ui_testing.automation.reporting.allure_helpers import attach_file, attach_image
except Exception:
    attach_image = None  # type: ignore
    attach_file = None  # type: ignore
try:
    from ui_testing.automation.state_snapshots import validate_exports
except Exception:
    validate_exports = None  # type: ignore
from ui_testing.tools.calibration import (
    capture_window_anchor,
    compute_offset,
    load_profile,
)
try:
    from ui_testing.automation.locator import (
        LocatorService,
        ManifestIndex,
        is_generic_automation_id,
        normalize_manifest,
    )
except Exception:
    try:
        from .locator import LocatorService, ManifestIndex, is_generic_automation_id, normalize_manifest  # type: ignore
    except Exception:
        @dataclass  # type: ignore[misc]
        class ManifestIndex:  # type: ignore[no-redef]
            groups: Dict[str, Dict[str, Dict[str, Any]]] = None  # type: ignore[assignment]
            lookup: Dict[str, Tuple[str, str]] = None  # type: ignore[assignment]

            def __post_init__(self) -> None:
                if self.groups is None:
                    self.groups = {}
                if self.lookup is None:
                    self.lookup = {}

            def get(self, automation_id: str):
                return self.lookup.get(str(automation_id))

            def contains(self, automation_id: str) -> bool:
                return str(automation_id) in self.lookup

        def normalize_manifest(manifest):  # type: ignore
            instance = ManifestIndex()
            if isinstance(manifest, dict):
                instance.groups = manifest  # type: ignore[assignment]
                instance.lookup = {}
                for group, mapping in manifest.items():
                    if isinstance(mapping, dict):
                        for name, payload in mapping.items():
                            auto_id = None
                            if isinstance(payload, dict):
                                auto_id = payload.get("automation_id") or payload.get("id")
                            else:
                                auto_id = str(payload)
                            if auto_id:
                                instance.lookup[str(auto_id)] = (str(group), str(name))
            return instance

        def is_generic_automation_id(value):  # type: ignore
            if not value:
                return True
            lowered = str(value).strip().lower()
            return lowered in {"", "window", "pane", "mainwindowcontrol"}

        class LocatorService:  # type: ignore[no-redef]
            def __init__(self, manifest=None) -> None:
                self.update_manifest(manifest)

            def update_manifest(self, manifest) -> None:
                self._index = normalize_manifest(manifest or {})

            @property
            def manifest(self):
                return getattr(self._index, "groups", {})

            @property
            def lookup(self):
                return getattr(self._index, "lookup", {})

            def contains(self, automation_id):
                if automation_id is None:
                    return False
                return str(automation_id) in self.lookup

            def manifest_entry(self, automation_id):
                if automation_id is None:
                    return None
                key = getattr(self._index, "lookup", {}).get(str(automation_id))
                if not key:
                    return None
                group, name = key
                return self.manifest.get(group, {}).get(name)

            def semantic_metadata(self, automation_id, control_type=None, registry=None):
                if not automation_id or is_generic_automation_id(automation_id):
                    return None
                payload = {"automation_id": str(automation_id)}
                if control_type:
                    payload["control_type"] = control_type
                if registry is not None and hasattr(registry, "get"):
                    try:
                        entry = registry.get(str(automation_id))
                    except Exception:
                        entry = None
                    if entry is not None:
                        payload["group"] = getattr(entry, "group", None)
                        payload["name"] = getattr(entry, "name", None)
                        ctrl = getattr(entry, "control_type", None)
                        if ctrl and "control_type" not in payload:
                            payload["control_type"] = ctrl
                        desc = getattr(entry, "description", None)
                        if desc:
                            payload["description"] = desc
                        return payload
                lookup_entry = getattr(self._index, "lookup", {}).get(str(automation_id))
                if lookup_entry:
                    group, name = lookup_entry
                    payload["group"] = group
                    payload["name"] = name
                    meta = self.manifest.get(group, {}).get(name, {})
                    desc = meta.get("description")
                    if desc:
                        payload["description"] = desc
                    ctrl = meta.get("control_type")
                    if ctrl and "control_type" not in payload:
                        payload["control_type"] = ctrl
                return payload
logger = logging.getLogger(__name__)

@dataclass
class PlayerConfig:
    scripts_dir: Path
    images_dir: Path
    results_dir: Path
    taskbar_crop_px: int = 60
    wait_between_actions: float = 1
    app_title_regex: Optional[str] = None  # regex to scope UIA search to app window
    diff_tolerance: float = 0.01  # 0.0 = exact; >0 allows small diffs
    diff_tolerance_percent: float = 0.01
    use_default_delay_always: bool = False
    use_automation_ids: bool = True
    use_screenshots: bool = True
    prefer_semantic_scripts: bool = True
    automation_manifest: Optional[Dict[str, Dict[str, Dict[str, Any]]]] = None
    use_ssim: bool = False
    ssim_threshold: float = 0.99
    automation_backend: str = "uia"
    appium_server_url: Optional[str] = None
    appium_capabilities: Optional[Dict[str, Any]] = None
    enable_allure: bool = True
    flake_stats_path: Optional[Path] = None
    state_snapshot_dir: Optional[Path] = None
    semantic_wait_timeout: float = 1.0
    semantic_poll_interval: float = 0.05
    calibration_profile: Optional[str] = None
    calibration_dir: Optional[Path] = None
    window_spec: Optional[WindowSpec] = None


class Player:
    def __init__(self, config: PlayerConfig) -> None:
        self.config = config
        if getattr(self.config, "window_spec", None) is None and DEFAULT_WINDOW_SPEC is not None:
            self.config.window_spec = DEFAULT_WINDOW_SPEC
        self._locator = LocatorService(config.automation_manifest or {})
        self.update_automation_manifest(config.automation_manifest or {})
        self._calibration_profile = None
        self._current_anchor: Optional[Tuple[int, int]] = None
        self._calibration_offset: Tuple[int, int] = (0, 0)
        self._initialize_calibration()

        self._screenshot_comparator = ScreenshotComparator(
            use_ssim=getattr(config, "use_ssim", False),
            ssim_threshold=getattr(config, "ssim_threshold", 0.99),
        )
        if getattr(self.config, "use_ssim", False) and not self._screenshot_comparator.using_ssim:
            logger.warning(
                "SSIM comparisons requested but scikit-image is not available. Disabling SSIM for this session."
            )
            self.config.use_ssim = False

        self._stop_event = threading.Event()
        self._flake_tracker = None
        if FlakeTracker is not None and getattr(config, "flake_stats_path", None):
            self._flake_tracker = FlakeTracker(Path(config.flake_stats_path))
        self._semantic_context: Optional[SemanticContext] = None
        self._semantic_disabled = False
        self._semantic_mode_active: bool = False
        self._metrics = PlaybackMetrics()
        self._semantic_registry_cache = None
        self._allure_enabled = bool(getattr(config, "enable_allure", True) and attach_image is not None)
        self._state_snapshot_dir = Path(config.state_snapshot_dir) if getattr(config, "state_snapshot_dir", None) else None
        self._current_script: Optional[str] = None
        self._uia_warning_logged = False
        self._window_log_once = False
        self._held_buttons: Set[str] = set()
        self._last_mouse_down_pos: Optional[Tuple[int, int]] = None
        self._semantic_wait_timeout = max(0.0, float(getattr(config, "semantic_wait_timeout", 1.0)))
        self._semantic_poll_interval = max(0.01, float(getattr(config, "semantic_poll_interval", 0.05)))

        # Lazy Explorer automation helper (wired by upcoming feature work)

    @property
    def ssim_available(self) -> bool:
        """Return True when SSIM comparisons can run (scikit-image is installed)."""
        return self._screenshot_comparator.ssim_available

    def update_automation_manifest(
        self, manifest: Optional[Dict[str, Dict[str, Dict[str, Any]]]]
    ) -> None:
        self._locator.update_manifest(manifest or {})
        structured = self._locator.manifest
        self.automation_manifest = structured
        self.config.automation_manifest = structured

        self._explorer_controller = None

        self._primary_bounds = self._init_primary_bounds()

        try:
            pyautogui.PAUSE = 0

            pyautogui.MINIMUM_DURATION = 0

        except Exception:
            pass

    def _initialize_calibration(self) -> None:
        profile_name = getattr(self.config, "calibration_profile", None)
        base_dir = getattr(self.config, "calibration_dir", None)
        if not profile_name or base_dir is None:
            self._calibration_profile = None
            self._current_anchor = None
            self._calibration_offset = (0, 0)
            return
        profile = load_profile(base_dir, profile_name)
        if profile is None:
            logger.debug("Calibration profile '%s' not found in %s", profile_name, base_dir)
            self._calibration_profile = None
            self._current_anchor = None
            self._calibration_offset = (0, 0)
            return
        anchor = capture_window_anchor(getattr(self.config, "window_spec", None))
        if anchor is None:
            logger.warning("Calibration profile '%s' requested but ENFIRE window anchor unavailable.", profile_name)
            self._calibration_profile = None
            self._current_anchor = None
            self._calibration_offset = (0, 0)
            return
        dx, dy = compute_offset(profile, (anchor[0], anchor[1]))
        self._calibration_profile = profile
        self._current_anchor = (int(anchor[0]), int(anchor[1]))
        self._calibration_offset = (dx, dy)
        if dx or dy:
            logger.info("Applied calibration offset (%s, %s) using profile '%s'", dx, dy, profile.name)

    def _calibrated_point(self, x: int, y: int) -> Tuple[int, int]:
        dx, dy = self._calibration_offset
        if dx == 0 and dy == 0:
            return x, y
        return x + dx, y + dy

    def _resolve_point(self, action: Optional[Dict[str, Any]], x: int, y: int) -> Tuple[int, int]:
        rel_x = self._extract_action_value(action, "rel_x")
        rel_y = self._extract_action_value(action, "rel_y")
        if rel_x is not None and rel_y is not None and self._current_anchor is not None:
            try:
                return int(self._current_anchor[0]) + int(rel_x), int(self._current_anchor[1]) + int(rel_y)
            except Exception:
                pass
        return self._calibrated_point(x, y)

    def _resolve_path(
        self, action: Optional[Dict[str, Any]], coords: List[Tuple[int, int]]
    ) -> List[Tuple[int, int]]:
        rel_path = self._extract_action_value(action, "rel_path")
        if rel_path and self._current_anchor is not None:
            anchor_x, anchor_y = self._current_anchor
            resolved: List[Tuple[int, int]] = []
            for point in rel_path:
                try:
                    px, py = int(point[0]), int(point[1])
                except Exception:
                    continue
                resolved.append((anchor_x + px, anchor_y + py))
            if resolved:
                return resolved
        return [self._calibrated_point(px, py) for px, py in coords]

    @staticmethod
    def _extract_action_value(action: Optional[Dict[str, Any]], key: str):
        if action is None:
            return None
        if isinstance(action, dict):
            return action.get(key)
        return getattr(action, key, None)

    def _semantic_context_kwargs(self) -> Dict[str, Any]:
        kwargs: Dict[str, Any] = {
            "backend": getattr(self.config, "automation_backend", "uia"),
            "appium_server_url": getattr(self.config, "appium_server_url", None),
            "appium_capabilities": getattr(self.config, "appium_capabilities", None) or None,
        }
        try:
            if WindowSpec is not None:
                spec = WindowSpec(
                    title_regex=self._normalized_title_regex(),
                    class_name=getattr(DEFAULT_WINDOW_SPEC, "class_name", None),
                )
                kwargs["window_spec"] = spec
        except Exception:
            pass
        return kwargs

    def _normalized_title_regex(self) -> Optional[str]:
        raw = getattr(self.config, "app_title_regex", None)
        if raw is None or str(raw).strip() == "":
            default = getattr(DEFAULT_WINDOW_SPEC, "title_regex", None)
            return default
        value = str(raw).strip()
        if not value:
            return getattr(DEFAULT_WINDOW_SPEC, "title_regex", None)
        # If it already looks like a regex (contains meta characters), trust it;
        # otherwise escape and wrap so partial titles still match.
        if re.search(r"[.^$*+\[\]|()?]", value):
            return value
        return f".*{re.escape(value)}.*"

    def _semantic_session(self) -> Optional[AutomationSession]:
        if not getattr(self.config, "prefer_semantic_scripts", True):
            return None
        if not getattr(self.config, "use_automation_ids", True):
            return None
        if SemanticContext is None or AutomationSession is None:
            return None
        if self._semantic_disabled:
            return None
        try:
            self._ensure_app_window()
            if self._semantic_context is None:
                self._semantic_context = get_semantic_context(**self._semantic_context_kwargs())
                self._semantic_registry_cache = None
            return self._semantic_context.session
        except PywinautoUnavailableError as exc:
            logger.warning(
                "Semantic automation unavailable: %s. Install pywinauto (pip install pywinauto) "
                "and ensure UI Testing runs with the same elevation as ENFIRE.",
                exc,
            )
        except Exception as exc:
            logger.warning(
                "Semantic session attach failed (%s). Verify ENFIRE is running and run UI Testing with matching privileges "
                "(Run as administrator if ENFIRE was launched elevated).",
                exc,
            )
            self._log_available_windows(self._normalized_title_regex() or "<unset>")
        self._semantic_context = None
        self._semantic_registry_cache = None
        return None

    def _semantic_registry(self):
        if self._semantic_disabled or SemanticContext is None:
            return None
        if self._semantic_registry_cache is not None:
            return self._semantic_registry_cache
        ctx = self._semantic_context
        if ctx is None:
            try:
                ctx = get_semantic_context(**self._semantic_context_kwargs())
                self._semantic_context = ctx
            except Exception as exc:
                logger.debug("Unable to resolve semantic context for templates: %s", exc)
                return None
        try:
            registry = ctx.registry
        except Exception as exc:
            logger.debug("Semantic registry unavailable: %s", exc)
            self._semantic_context = None
            self._semantic_registry_cache = None
            return None
        self._semantic_registry_cache = registry
        return registry

    def _run_semantic_template(self, semantic_meta: Dict[str, Any], expected: Any) -> None:
        if not self._semantic_mode_active:
            return
        group = semantic_meta.get("group")
        name = semantic_meta.get("name")
        if not group or not name:
            return
        ctx = self._semantic_context
        if ctx is None:
            try:
                ctx = get_semantic_context(**self._semantic_context_kwargs())
                self._semantic_context = ctx
            except Exception as exc:
                logger.debug("Unable to resolve semantic context for templates: %s", exc)
                return
        screen = ctx.resolve_screen_for_group(group)
        if screen is None:
            logger.warning("Semantic template skipped: no screen mapped for %s.%s", group, name)
            return
        expected_str = "" if expected is None else str(expected)
        identifier = f"semantic:{group}.{name}"
        try:
            handled = False
            if group == "BridgeIds" and name == "BridgeMlcResults":
                screen.assert_mlc(expected_str)
                handled = True
            elif group == "TerrainIds" and name == "TerrainName":
                screen.assert_name(expected_str)
                handled = True
            if not handled:
                logger.debug("Semantic template has no explicit handler for %s.%s", group, name)
        except AssertionError as exc:
            message = f"Semantic template assertion failed for {identifier}: {exc}"
            logger.warning(message)
            self._record_failure(identifier)
            raise AssertionError(message) from exc
        except Exception as exc:
            message = f"Semantic template execution failed for {identifier}: {exc}"
            logger.warning(message)
            self._record_failure(identifier)
            raise AssertionError(message) from exc

    def _record_failure(self, identifier: str) -> None:
        if self._flake_tracker and self._current_script:
            try:
                self._flake_tracker.record_failure(self._current_script, identifier)
            except Exception:
                pass

    def _log_uia_hint(self, exc: Exception) -> None:
        if self._uia_warning_logged:
            return
        message = str(exc)
        matched = False
        if ElementNotFoundError is not None and isinstance(exc, ElementNotFoundError):
            matched = True
        elif WindowNotFoundError is not None and isinstance(exc, WindowNotFoundError):
            matched = True
        elif "Access is denied" in message:
            matched = True
        if matched:
            logger.info(
                "UIA fallback to coordinates detected. If ENFIRE was launched from Visual Studio or elevated, "
                "launch UI Testing with matching privileges (Run as administrator) so UI Automation can reach the controls."
            )
            self._uia_warning_logged = True

    def _ensure_app_window(self) -> None:
        if Desktop is None:
            return
        regex = self._normalized_title_regex()
        if not regex:
            return
        for _ in range(6):
            try:
                Desktop(backend="uia").window(title_re=regex).wait("exists ready", timeout=0.5)
                return
            except Exception:
                time.sleep(0.5)
        self._log_available_windows(regex)

    def _log_available_windows(self, regex: str) -> None:
        if self._window_log_once or Desktop is None:
            return
        self._window_log_once = True
        try:
            titles = [w.window_text() for w in Desktop(backend="uia").windows() if w.window_text()]
        except Exception:
            titles = []
        if titles:
            logger.info("UIA visible windows: %s", titles)
        else:
            logger.info("UIA reported no visible top-level windows when matching regex %s", regex)

    def _attach_flake_stats_artifact(self) -> None:
        if not self._allure_enabled or attach_file is None:
            return
        if not self._flake_tracker:
            return
        path = getattr(self._flake_tracker, "path", None)
        if not path:
            return
        try:
            path_obj = Path(path)
        except Exception:
            return
        if not path_obj.exists():
            return
        try:
            attach_file("flake-stats.json", path_obj, attachment_type="application/json")
        except Exception:
            pass

    def _run_state_snapshot_checks(self) -> None:
        if self._state_snapshot_dir and validate_exports is not None:
            try:
                validate_exports(self._state_snapshot_dir)
            except AssertionError as exc:
                logger.warning("State snapshot validation failed: %s", exc)
                self._record_failure("state_snapshot")
                raise
            except Exception as exc:
                logger.debug("State snapshot validation skipped: %s", exc)

    def play(self, script_name: str) -> List[Dict[str, Any]]:
        """Returns per-checkpoint results and writes an Excel summary per run."""

        script_path = self._select_script_path(script_name)

        with script_path.open("r", encoding="utf-8") as f:
            actions: List[Dict[str, Any]] = json.load(f)

        results: List[Dict[str, Any]] = []
        self._metrics.reset()
        self._current_script = script_name
        self._held_buttons.clear()
        semantic_mode = (
            bool(getattr(self.config, "use_automation_ids", True))
            and bool(getattr(self.config, "prefer_semantic_scripts", True))
            and Desktop is not None
        )
        self._semantic_mode_active = semantic_mode
        try:
            assert_count = 0
            screenshot_count = 0

            shot_idx = 0
        
            base_code = dotted_code_from_test_name(Path(script_name).name)
        
            total_actions = len(actions)
        
            i = 0
        
            while i < total_actions:
                action = actions[i]
        
                a_type = action.get("action_type")
        
                pre = self._compute_action_delay(a_type, action)
        
                if pre > 0:
                    remaining = pre
        
                    while remaining > 0 and not self.should_stop():
                        chunk = min(0.1, remaining)
        
                        time.sleep(chunk)
        
                        remaining -= chunk
        
                if self.should_stop():
                    break
        
                if a_type and str(a_type).startswith("explorer."):
                    self._play_explorer_action(action)
        
                    i += 1
        
                    continue
        
                if a_type == "click":
                    x, y = int(action["x"]), int(action["y"])

                    auto_id: Optional[str] = action.get("auto_id")
                    ctrl_type: Optional[str] = action.get("control_type")

                    if not self._in_primary_monitor(x, y):
                        logger.info(f"Playback: click skipped outside primary monitor at ({x}, {y})")

                        i += 1

                        continue

                    use_uia = (
                        semantic_mode
                        and auto_id
                        and not is_generic_automation_id(auto_id)
                    )
                    if use_uia and not self._locator.contains(auto_id):
                        logger.debug("AutomationId %s not found in manifest; falling back to coordinates.", auto_id)
                        use_uia = False

                    property_filter: Optional[Tuple[str, Any]] = None
                    if semantic_mode:
                        prop_name = (
                            action.get("property")
                            or action.get("property_name")
                            or action.get("propertyName")
                        )
                        prop_expected = action.get("expected")
                        if prop_name and prop_expected not in (None, ""):
                            property_filter = (str(prop_name), prop_expected)

                    if use_uia:
                        target = None
                        mode = None
                        session = self._semantic_session()
                        if session is not None:
                            try:
                                candidate = session.resolve_control(automation_id=str(auto_id), control_type=ctrl_type)
                                if candidate is not None and self._match_property(candidate, property_filter):
                                    target = candidate
                                    mode = "semantic"
                                else:
                                    logger.debug(
                                        "Semantic session candidate for auto_id='%s' did not match property filter.",
                                        auto_id,
                                    )
                            except Exception as exc:
                                logger.debug("Semantic session click failed: %s", exc)
                        if target is None:
                            target = self._resolve_element_by_auto_id(
                                str(auto_id),
                                ctrl_type,
                                property_filter,
                                skip_semantic=True,
                            )
                            if target is not None:
                                mode = "uia"
                        if target is not None:
                            try:
                                target.click_input()
                                if mode == "semantic":
                                    logger.info(
                                        f"Playback(Semantic): click auto_id='{auto_id}'"
                                        f"{' ctrl=' + ctrl_type if ctrl_type else ''}"
                                    )
                                    self._metrics.note_click("semantic", auto_id, ctrl_type, (x, y))
                                else:
                                    logger.info(
                                        f"Playback(UIA): click auto_id='{auto_id}'"
                                        f"{' ctrl=' + ctrl_type if ctrl_type else ''}"
                                    )
                                    self._metrics.note_click("uia", auto_id, ctrl_type, (x, y))
                                if self._semantic_mode_active and property_filter:
                                    self._wait_for_property(str(auto_id), ctrl_type, property_filter)
                                time.sleep(0.005)
                                i += 1
                                continue
                            except Exception as e:
                                logger.warning(
                                    f"UIA click failed for auto_id='{auto_id}' (fallback to coords): {e}"
                                )
                                self._log_available_windows(self._normalized_title_regex() or "<unspecified>")
                                self._log_uia_hint(e)

                    cx, cy = self._resolve_point(action, x, y)
                    fallback_note = "" if auto_id else " [coordinate fallback]"
                    if (cx, cy) != (x, y):
                        logger.info(f"Playback: click at ({x}, {y}) -> calibrated ({cx}, {cy}){fallback_note}")
                    else:
                        logger.info(f"Playback: click at ({x}, {y}){fallback_note}")

                    pyautogui.click(cx, cy, _pause=False)
                    self._metrics.note_click("coordinate", auto_id, ctrl_type, (cx, cy))
        
                elif a_type == "mouse_down":
                    x = int(action.get("x", 0))
        
                    y = int(action.get("y", 0))
        
                    button = str(action.get("button") or "left").lower()

                    cx, cy = self._resolve_point(action, x, y)
                    if (cx, cy) != (x, y):
                        logger.debug(f"Playback: mouse_down({button}) at ({x}, {y}) -> ({cx}, {cy})")
                    else:
                        logger.debug(f"Playback: mouse_down({button}) at ({x}, {y})")

                    try:
                        pyautogui.mouseDown(x=cx, y=cy, button=button, _pause=False)
                        self._held_buttons.add(button)
                        self._last_mouse_down_pos = (cx, cy)

                    except Exception as exc:
                        logger.warning(f"mouse_down failed at ({cx}, {cy}): {exc}")
        
                elif a_type == "mouse_move":
                    button = action.get("button")
        
                    try:
                        duration_val = action.get("move_duration")
                        move_duration = max(0.0, float(duration_val)) if duration_val is not None else 0.0
                    except Exception:
                        move_duration = 0.0

                    if button:
                        coords: List[Tuple[int, int]] = []
                        coord_actions: List[Dict[str, Any]] = []
                        j = i
                        raw_duration = 0.0
                        while j < total_actions:
                            next_action = actions[j]
                            if next_action.get("action_type") != "mouse_move" or next_action.get("button") != button:
                                break
                            coords.append((int(next_action.get("x", 0)), int(next_action.get("y", 0))))
                            coord_actions.append(next_action)
                            try:
                                delay_component = next_action.get("delay", 0.0) or 0.0
                                raw_duration += max(float(delay_component), 0.0)
                            except Exception:
                                pass
                            j += 1

                        filtered_coords: List[Tuple[int, int]] = []
                        filtered_actions: List[Optional[Dict[str, Any]]] = []
                        for idx, (px, py) in enumerate(coords):
                            if self._in_primary_monitor(px, py):
                                filtered_coords.append((px, py))
                                filtered_actions.append(coord_actions[idx] if idx < len(coord_actions) else None)
                        coords = filtered_coords
                        coord_actions = filtered_actions
                        resolved_coords: List[Tuple[int, int]] = []
                        for idx, (px, py) in enumerate(coords):
                            source_action = coord_actions[idx] if idx < len(coord_actions) else None
                            resolved_coords.append(self._resolve_point(source_action, px, py))
                        coords = resolved_coords

                        if len(coords) > 1:
                            start_pos = self._last_mouse_down_pos
                            if (
                                start_pos is not None
                                and self._in_primary_monitor(start_pos[0], start_pos[1])
                                and coords[0] != start_pos
                            ):
                                coords.insert(0, start_pos)
                            total_duration = raw_duration if raw_duration > 0 else None
                            self._play_drag_path(coords, button, total_duration)
                            self._metrics.note_drag(button, len(coords))
                            i = j
                            continue
                        else:
                            # Fallback to a simple move if no usable path
                            if coords:
                                try:
                                    px, py = coords[-1]
                                    pyautogui.moveTo(px, py, duration=move_duration, _pause=False)
                                except Exception as exc:
                                    logger.warning(f"mouse_move failed to ({coords[-1][0]}, {coords[-1][1]}): {exc}")
                            i = j
                            continue

                    x = int(action.get("x", 0))
                    y = int(action.get("y", 0))

                    try:
                        cx, cy = self._resolve_point(action, x, y)
                        if move_duration > 0:
                            pyautogui.moveTo(cx, cy, duration=move_duration, _pause=False)
                        else:
                            pyautogui.moveTo(cx, cy, _pause=False)
                    except Exception as exc:
                        logger.warning(f"mouse_move failed to ({cx}, {cy}): {exc}")

                elif a_type == "mouse_up":
                    x = int(action.get("x", 0))
        
                    y = int(action.get("y", 0))
        
                    button = str(action.get("button") or "left").lower()
        
                    cx, cy = self._resolve_point(action, x, y)
                    if (cx, cy) != (x, y):
                        logger.debug(f"Playback: mouse_up({button}) at ({x}, {y}) -> ({cx}, {cy})")
                    else:
                        logger.debug(f"Playback: mouse_up({button}) at ({x}, {y})")

                    try:
                        pyautogui.mouseUp(x=cx, y=cy, button=button, _pause=False)
                        self._held_buttons.discard(button)

                    except Exception as exc:
                        logger.warning(f"mouse_up failed at ({cx}, {cy}): {exc}")
                    finally:
                        self._last_mouse_down_pos = None
        
                elif a_type == "drag":
                    raw_path = action.get("path") or []
        
                    button = str(action.get("button") or "left").lower()
        
                    coords: List[Tuple[int, int]] = []
                    for point in raw_path:
                        try:
                            px, py = int(point[0]), int(point[1])
                        except Exception:
                            continue
                        coords.append((px, py))
                    coords = [pt for pt in coords if self._in_primary_monitor(pt[0], pt[1])]
                    rel_path = self._extract_action_value(action, "rel_path")
                    if rel_path and self._current_anchor is not None:
                        anchor_x, anchor_y = self._current_anchor
                        resolved: List[Tuple[int, int]] = []
                        for point in rel_path:
                            try:
                                px, py = int(point[0]), int(point[1])
                            except Exception:
                                continue
                            resolved.append((anchor_x + px, anchor_y + py))
                        if resolved:
                            coords = resolved
                        else:
                            coords = [self._calibrated_point(pt[0], pt[1]) for pt in coords]
                    else:
                        coords = [self._calibrated_point(pt[0], pt[1]) for pt in coords]
        
                    if len(coords) > 1:
                        logger.debug(f"Playback: drag path ({len(coords)} points) [{button}]")

                        drag_duration = action.get("drag_duration")
                        try:
                            drag_duration_val: Optional[float] = (
                                float(drag_duration) if drag_duration is not None else None
                            )
                        except Exception:
                            drag_duration_val = None
                        self._play_drag_path(coords, button, drag_duration_val)
                        self._metrics.note_drag(button, len(coords))

                    else:
                        logger.debug("Playback: drag skipped (insufficient path points)")

                elif a_type == "key":
                    key_name = action.get("key")
        
                    if key_name:
                        logger.info(f"Playback: key {key_name}")
        
                        try:
                            pyautogui.press(key_name)
        
                        except Exception as exc:
                            logger.warning(f"key press failed ({key_name}): {exc}")
        
                elif a_type == "hotkey":
                    keys = action.get("keys")
        
                    if not keys:
                        i += 1
        
                        continue
        
                    if not isinstance(keys, list):
                        try:
                            keys = list(keys)
        
                        except Exception:
                            keys = [str(keys)]
        
                    str_keys = [str(k).lower() for k in keys]
        
                    label = ' + '.join(str_keys)
        
                    logger.info(f"Playback: hotkey {label}")
                    self._play_hotkey(str_keys)
        
                elif a_type == "scroll":
                    x = action.get("x")
        
                    y = action.get("y")
        
                    dx = int(action.get("scroll_dx", 0) or 0)
        
                    dy = int(action.get("scroll_dy", 0) or 0)
        
                    if dx or dy:
                        if x is not None and y is not None:
                            xi, yi = self._resolve_point(action, int(x), int(y))
                            try:
                                cur = pyautogui.position()
                                cx = int(getattr(cur, 'x', cur[0]))
                                cy = int(getattr(cur, 'y', cur[1]))
                            except Exception:
                                cx = cy = None
                            try:
                                if cx != xi or cy != yi:
                                    pyautogui.moveTo(xi, yi, duration=0)
                            except Exception:
                                pass

                        else:
                            xi = yi = None

                        if xi is not None and yi is not None and not self._in_primary_monitor(xi, yi):
                            logger.info(f"Playback: scroll ignored outside primary monitor at ({xi}, {yi})")
        
                            i += 1
        
                            continue
        
                        logger.info(f"Playback: scroll at ({xi if xi is not None else '?'}, {yi if yi is not None else '?'}) dx={dx}, dy={dy}")
        
                        try:
                            if dy:
                                pyautogui.scroll(dy, x=xi, y=yi)
        
                            if dx:
                                if hasattr(pyautogui, 'hscroll'):
                                    pyautogui.hscroll(dx, x=xi, y=yi)
        
                                else:
                                    logger.debug("Horizontal scroll not supported on this platform")
        
                            time.sleep(0.005)
        
                        except Exception as exc:
                            logger.warning(f"scroll failed at ({xi if xi is not None else '?'}, {yi if yi is not None else '?'}):: {exc}")
        
                elif a_type == "type":
                    text = action.get("text", "")
        
                    safe_preview = text.replace("\n", "<ENTER>")
        
                    logger.info(f"Playback: type '{safe_preview}'")
        
                    pyautogui.typewrite(text, interval=0.02)
        
                elif a_type == "assert.property":
                    if not getattr(self.config, "prefer_semantic_scripts", True):
                        logger.debug("Playback: semantic assertion skipped (semantic checks disabled).")
                    elif not getattr(self.config, "use_automation_ids", True):
                        logger.debug("Playback: semantic assertion skipped (automation IDs disabled).")
                    elif Desktop is None:
                        logger.debug("Playback: semantic assertion skipped (UI Automation backend unavailable).")
                    else:
                        before_len = len(results)
                        self._handle_assert_property(action, results)
                        if len(results) > before_len:
                            assert_count += 1
                elif a_type == "screenshot":
                    if not getattr(self.config, "use_screenshots", True):
                        logger.info("Playback: screenshot checkpoint skipped (screenshots disabled)")
                        i += 1
                        continue
                    prev_pos = None
        
                    try:
                        pos = pyautogui.position()
        
                        if hasattr(pos, "x") and hasattr(pos, "y"):
                            prev_pos = (int(pos.x), int(pos.y))
        
                        else:
                            prev_pos = (int(pos[0]), int(pos[1]))
        
                    except Exception:
                        prev_pos = None
        
                    logger.info(f"Playback: screenshot #{shot_idx}")
                    time.sleep(0.5)
                    test_img = self._capture_screenshot_primary()
        
                    if prev_pos is not None:
                        try:
                            pyautogui.moveTo(prev_pos[0], prev_pos[1], duration=0)
        
                        except Exception:
                            pass
        
                    img_dir = self.config.images_dir / script_name
        
                    img_dir.mkdir(parents=True, exist_ok=True)
        
                    test_name = ensure_png_name(0, shot_idx, "T")
        
                    test_path = img_dir / test_name
        
                    test_img.save(test_path)
        
                    orig_name = ensure_png_name(0, shot_idx, "O")
        
                    orig_path = img_dir / orig_name
        
                    diff_tolerance = self._diff_tolerance()
                    comparison = self._screenshot_comparator.compare(orig_path, test_path, diff_tolerance)
                    passed = comparison.passed
                    diff_pct = comparison.diff_percent
                    diff_path = comparison.diff_path
                    highlight_path = comparison.highlight_path
                    ssim_score = comparison.ssim_score
                    ssim_threshold_value = comparison.ssim_threshold
                    identifier = f"screenshot:{shot_idx}"

                    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                    metrics_parts = [f"Δ {float(diff_pct):.3f}%"]
                    if ssim_score is not None:
                        metrics_parts.append(f"SSIM {ssim_score:.4f}")

                    result = {

                        "index": shot_idx,

                        "original": str(orig_path),
        
                        "test": str(test_path),
        
                        "diff_percent": round(float(diff_pct), 3),
                        "pixel_diff_percent": round(float(diff_pct), 3),
                        "metrics": " | ".join(metrics_parts),
                        "ssim_score": round(float(ssim_score), 4) if ssim_score is not None else None,
                        "ssim_threshold": float(ssim_threshold_value) if ssim_threshold_value is not None else None,

                        "status": "pass" if passed else "fail",

                        "timestamp": timestamp,

                    }

                    logger.info(f"Result: screenshot #{shot_idx} -> {'PASS' if passed else 'FAIL'}")

                    results.append(result)
                    if result["status"] != "pass":
                        self._record_failure(identifier)
                        if self._allure_enabled and attach_image is not None:
                            if test_path.exists():
                                attach_image("Screenshot (test)", test_path)
                            if orig_path.exists():
                                attach_image("Screenshot (original)", orig_path)
                            if diff_path and diff_path.exists():
                                attach_image("Diff (D)", diff_path)
                            if highlight_path and highlight_path.exists():
                                attach_image("Diff (H)", highlight_path)
        
                    shot_idx += 1
                    screenshot_count += 1
        
                i += 1
        
            summary_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            validation_fail = any(r.get("status") == "fail" for r in results)
            validation_total = assert_count + screenshot_count
            note_parts = [f"Asserts: {assert_count}", f"Screenshots: {screenshot_count}"]
            click_counts = self._metrics.click_counts
            note_parts.append(
                "Clicks semantic/UIA/coords: "
                f"{click_counts.get('semantic', 0)}/"
                f"{click_counts.get('uia', 0)}/"
                f"{click_counts.get('coordinate', 0)}"
            )
            if click_counts.get("coordinate", 0):
                note_parts.append(
                    f"Coordinates used for {click_counts.get('coordinate', 0)} click(s)."
                )
            else:
                note_parts.append("All clicks resolved via AutomationIds (semantic/UIA).")
            if self._metrics.drag_count:
                note_parts.append(f"Drags replayed: {self._metrics.drag_count} (coordinate path)")
            if validation_total == 0:
                summary_status = "warn"
                note_parts.append("No semantic assertions or screenshot checkpoints executed.")
                logger.warning("Playback summary [%s]: no validations executed (asserts=0, screenshots=0).", script_name)
            elif validation_fail:
                summary_status = "fail"
                note_parts.append("At least one validation failed.")
            else:
                summary_status = "pass"
                note_parts.append("All validations passed.")
            summary_entry = {
                "index": "summary",
                "timestamp": summary_timestamp,
                "original": "",
                "test": "",
                "diff_percent": "",
                "status": summary_status,
                "note": " | ".join(note_parts),
                "assertions": assert_count,
                "screenshots": screenshot_count,
            }
            results.append(summary_entry)

            if self._metrics.click_history:
                logger.info("Click playback modes: %s", "; ".join(self._metrics.click_history))
            if self._metrics.drag_history:
                logger.info("Drag playback details: %s", "; ".join(self._metrics.drag_history))
            if click_counts.get("uia", 0) or click_counts.get("semantic", 0):
                logger.info(
                    "Non-coordinate modes handled %d clicks (semantic %d / UIA %d).",
                    click_counts.get("uia", 0) + click_counts.get("semantic", 0),
                    click_counts.get("semantic", 0),
                    click_counts.get("uia", 0),
                )

            self._run_state_snapshot_checks()
        finally:
            self._semantic_mode_active = False
            self._current_script = None
            self._held_buttons.clear()
            self._last_mouse_down_pos = None
        # Write Excel after this test

        self._write_excel_results(script_name, results)
        self._attach_flake_stats_artifact()

        return results

    def _handle_assert_property(self, action: Dict[str, Any], results: List[Dict[str, Any]]) -> None:
        if not self._semantic_mode_active:
            return
        if not getattr(self.config, "prefer_semantic_scripts", True):
            return
        if not getattr(self.config, "use_automation_ids", True):
            return
        semantic_meta = action.get("semantic") or {}
        auto_id = action.get("auto_id") or semantic_meta.get("automation_id")
        if not auto_id:
            logger.warning("assert.property skipped (missing auto_id)")
            return
        if is_generic_automation_id(auto_id):
            logger.debug("assert.property skipped for generic auto_id=%s", auto_id)
            return
        ctrl_type = action.get("control_type") or semantic_meta.get("control_type")
        if not self._locator.contains(auto_id):
            logger.debug("AutomationId %s not defined in manifest", auto_id)
            return
        prop_name = (
            action.get("property")
            or action.get("property_name")
            or action.get("propertyName")
            or "name"
        )
        comparator = str(action.get("compare") or action.get("comparison") or "equals").strip().lower()
        expected = action.get("expected")
        element = None
        property_filter: Optional[Tuple[str, Any]] = None
        if prop_name and expected not in (None, ""):
            property_filter = (str(prop_name), expected)
        session = self._semantic_session() if self._semantic_mode_active else None
        if session is not None:
            try:
                candidate = session.resolve_control(automation_id=str(auto_id), control_type=ctrl_type)
                if candidate is not None and self._match_property(candidate, property_filter):
                    element = candidate
                else:
                    logger.debug(
                        "Semantic resolve candidate for %s did not satisfy property filter.",
                        auto_id,
                    )
            except Exception as exc:
                logger.debug("Semantic resolve failed for %s: %s", auto_id, exc)
        if element is None and Desktop is not None:
            element = self._resolve_element_by_auto_id(
                str(auto_id),
                ctrl_type,
                property_filter,
                skip_semantic=True,
            )
        if element is None:
            logger.warning("assert.property failed: element auto_id='%s' not found", auto_id)
            self._record_assert_result(results, str(auto_id), prop_name, expected, None, False, "not found", semantic_meta if semantic_meta else None)
            return
        actual = self._read_element_property(element, prop_name)
        passed, note = self._compare_property(actual, expected, comparator)
        if semantic_meta:
            try:
                self._run_semantic_template(semantic_meta, expected)
            except AssertionError as exc:
                passed = False
                note = str(exc)
            except Exception as exc:
                logger.debug("Semantic template ignored: %s", exc)
        extra = ""
        group = semantic_meta.get("group")
        name = semantic_meta.get("name")
        if group and name:
            extra = f" [{group}.{name}]"
        logger.info(
            "Assert property: auto_id='%s'%s property='%s' comparator='%s' -> %s",
            auto_id,
            extra,
            prop_name,
            comparator,
            "PASS" if passed else "FAIL",
        )
        self._record_assert_result(results, str(auto_id), prop_name, expected, actual, passed, note, semantic_meta if semantic_meta else None)

    def _capture_screenshot_primary(self) -> Image.Image:
        prev_failsafe = pyautogui.FAILSAFE

        pyautogui.FAILSAFE = False

        try:
            sw, sh = pyautogui.size()

            pyautogui.moveTo(sw - 5, sh - 5, duration=0)

            shot = pyautogui.screenshot()

        finally:
            pyautogui.FAILSAFE = prev_failsafe

        if self.config.taskbar_crop_px > 0:
            w, h = shot.size

            shot = shot.crop((0, 0, w, h - self.config.taskbar_crop_px))

        return shot

    def _compare_exact(self, orig_path: Path, test_path: Path) -> bool:
        result = self._screenshot_comparator.compare(orig_path, test_path, self._diff_tolerance())
        return result.passed

    def _diff_tolerance(self) -> float:
        return float(
            getattr(
                self.config,
                "diff_tolerance_percent",
                getattr(self.config, "diff_tolerance", 0.0),
            )
        )

    def _split_hierarchy(self, script_name: str) -> Tuple[str, str, str]:
        p = Path(script_name)

        test = p.name

        sec = p.parent.name if p.parent.name else ""

        proc = p.parent.parent.name if p.parent.parent and p.parent.parent.name else ""

        return proc, sec, test

    def request_stop(self, clear_only: bool=False):
        """If clear_only=True, clears previous stop; else sets stop."""

        if clear_only:
            self._stop_event.clear()

        else:
            self._stop_event.set()

    def _compute_action_delay(self, action_type: str, action: Dict[str, Any]) -> float:
        default_wait = float(getattr(self.config, 'wait_between_actions', 0.0) or 0.0)

        use_default = bool(getattr(self.config, 'use_default_delay_always', False))

        drag_actions = {'mouse_down', 'mouse_move', 'mouse_up'}

        scroll_actions = {'scroll'}

        if use_default and action_type not in drag_actions and action_type not in scroll_actions:
            target = default_wait

        else:
            target = action.get('delay', 0.0 if action_type in scroll_actions else default_wait)

        try:
            delay = float(target)

        except Exception:
            if action_type in drag_actions:
                delay = 0.01

            elif action_type in scroll_actions:
                delay = 0.0

            else:
                delay = default_wait

        if delay < 0:
            delay = 0.0

        if use_default and action_type in drag_actions and delay == 0.0:
            delay = 0.01

        if action_type in drag_actions:
            delay *= 0.1
        if action_type == "mouse_up" and delay < 0.05:
            delay = 0.05

        return delay

    def _play_drag_path(
        self,
        coords: List[Tuple[int, int]],
        button: str,
        total_duration: Optional[float] = None,
    ) -> None:
        if len(coords) < 2:
            return

        sampled = self._downsample_points(coords)
        if len(sampled) < 2:
            return

        start_x, start_y = sampled[0]
        try:
            pyautogui.moveTo(start_x, start_y, duration=0, _pause=False)
        except Exception as exc:
            logger.debug("Initial drag move failed to (%s, %s): %s", start_x, start_y, exc)

        if button and button not in self._held_buttons:
            try:
                pyautogui.mouseDown(button=button, _pause=False)
                self._held_buttons.add(button)
                logger.debug("Supplemental mouse_down for drag using button '%s'", button)
            except Exception as exc:
                logger.debug("Supplemental mouse_down failed for drag: %s", exc)

        duration_value = 0.0
        if total_duration is not None:
            try:
                duration_value = max(float(total_duration), 0.0)
            except Exception:
                duration_value = 0.0

        steps = max(1, len(sampled) - 1)
        per_step = duration_value / steps if duration_value > 0 else 0.0
        if 0 < per_step < 0.005:
            per_step = 0.005

        for px, py in sampled[1:]:
            try:
                if per_step > 0:
                    pyautogui.moveTo(px, py, duration=per_step, _pause=False)
                else:
                    pyautogui.moveTo(px, py, duration=0, _pause=False)
            except Exception as exc:
                logger.warning(f"drag move failed to ({px}, {py}): {exc}")
                break

        # Give the UI a moment to register the selection before the next action (e.g., Ctrl+C)
        settle = per_step if per_step > 0 else 0.02
        time.sleep(min(0.05, max(settle, 0.02)))

    def _play_hotkey(self, keys: Sequence[str]) -> None:
        normalized = [self._normalize_hotkey_part(k) for k in keys if k]
        if not normalized:
            return
        primary = normalized[-1]
        modifiers = normalized[:-1]
        pressed_mods: List[str] = []
        try:
            for mod in modifiers:
                if not mod:
                    continue
                try:
                    pyautogui.keyDown(mod, _pause=False)
                    pressed_mods.append(mod)
                except Exception as exc:  # pragma: no cover - UI timing dependent
                    logger.warning("hotkey failed (keyDown %s): %s", mod, exc)
                time.sleep(0.02)
            if primary:
                try:
                    pyautogui.press(primary, _pause=False)
                except Exception as exc:
                    logger.warning("hotkey failed (press %s): %s", primary, exc)
            time.sleep(0.02)
        finally:
            for mod in reversed(pressed_mods):
                try:
                    pyautogui.keyUp(mod, _pause=False)
                except Exception as exc:
                    logger.debug("hotkey cleanup failed (keyUp %s): %s", mod, exc)
                time.sleep(0.01)

    def _wait_for_property(
        self,
        auto_id: str,
        ctrl_type: Optional[str],
        property_filter: Optional[Tuple[str, Any]],
    ) -> None:
        if not self._semantic_mode_active:
            return
        if not property_filter:
            return
        prop_name, expected = property_filter
        if not prop_name:
            return
        timeout = self._semantic_wait_timeout
        if timeout <= 0:
            return
        interval = min(self._semantic_poll_interval, timeout)
        if interval <= 0:
            interval = 0.01
        deadline = time.monotonic() + timeout
        while time.monotonic() < deadline:
            element = self._resolve_element_by_auto_id(
                auto_id,
                ctrl_type,
                property_filter,
                skip_semantic=True,
            )
            if element is not None and self._match_property(element, property_filter):
                return
            time.sleep(interval)
        logger.debug(
            "Semantic wait timed out for auto_id='%s' property='%s' expected='%s'",
            auto_id,
            prop_name,
            expected,
        )

    @staticmethod
    def _normalize_hotkey_part(key: str) -> str:
        value = str(key).strip().lower()
        if value in {"control", "ctl"}:
            return "ctrl"
        if value in {"windows", "command", "cmd"}:
            return "win"
        if value == "option":
            return "alt"
        return value

    def _downsample_points(self, coords: List[Tuple[int, int]], max_points: int = 120) -> List[Tuple[int, int]]:
        if len(coords) <= max_points:
            return coords

        stride = max(1, len(coords) // max_points)

        sampled = coords[::stride]

        if sampled[-1] != coords[-1]:
            sampled.append(coords[-1])

        return sampled

    def should_stop(self) -> bool:
        return self._stop_event.is_set()

    def _write_excel_results(self, script_name: str, results: List[Dict[str, Any]]) -> None:
        # Lazy import so users without Excel don't break other features

        try:
            from openpyxl import Workbook, load_workbook

            from openpyxl.styles import Font, Alignment, PatternFill

        except Exception as e:
            logger.warning(f"Excel export skipped (openpyxl not available): {e}")

            return

        proc, sec, test = self._split_hierarchy(script_name)

        sheet_name = (proc or "General")[:31] or "General"

        out_path = self.config.results_dir / "results_summary.xlsx"

        if out_path.exists():
            wb = load_workbook(out_path)

            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

            else:
                ws = wb.create_sheet(title=sheet_name)

                self._initialize_results_sheet(ws, Font, Alignment)

        else:
            wb = Workbook()

            ws = wb.active

            ws.title = sheet_name

            self._initialize_results_sheet(ws, Font, Alignment)

        existing_rows = ws.max_row

        headers_written = existing_rows > 1

        target_key = ((proc or ""), (sec or ""), (test or ""))

        if existing_rows > 1:
            rows_to_keep = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue

                row_key = ((row[0] or ""), (row[1] or ""), (row[2] or ""))

                if row_key != target_key:
                    rows_to_keep.append(list(row))

            if len(rows_to_keep) != existing_rows - 1:
                ws.delete_rows(2, existing_rows - 1)

                for values in rows_to_keep:
                    ws.append(values)

        for r in results:
            raw_index = r.get("index", 0)
            if isinstance(raw_index, str) and raw_index.lower() == "summary":
                continue

            try:
                checkpoint = f"{int(raw_index) + 1}"

            except Exception:
                checkpoint = str(raw_index)

            timestamp = r.get("timestamp", "")

            diff_value: Optional[float]

            try:
                diff_value = round(float(r.get("diff_percent")), 3)

            except Exception:
                diff_value = None

            status = "PASS" if r.get("status") == "pass" else "FAIL"

            row = [

                proc,

                sec,

                test,

                checkpoint,

                timestamp,

                diff_value,

                status,

                r.get("original", ""),

                r.get("test", ""),

            ]

            ws.append(row)

            last_row = ws.max_row

            # apply formatting

            timestamp_cell = ws.cell(row=last_row, column=5)

            timestamp_cell.alignment = Alignment(horizontal="center")

            diff_cell = ws.cell(row=last_row, column=6)

            if diff_value is None:
                diff_cell.value = None

            else:
                diff_cell.number_format = "0.000"

            result_cell = ws.cell(row=last_row, column=7)

            result_cell.alignment = Alignment(horizontal="center")

            if status == "PASS":
                result_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

            else:
                result_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        if not headers_written:
            self._apply_column_widths(ws)

        try:
            wb.save(out_path)

            logger.info(f"Excel results saved: {out_path}")
            if self._allure_enabled and attach_file is not None:
                try:
                    attach_file(
                        "results_summary.xlsx",
                        out_path,
                        attachment_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                except Exception:
                    pass

        except Exception as e:
            logger.warning(f"Failed to save Excel results: {e}")

    def _initialize_results_sheet(self, ws, Font, Alignment) -> None:
        headers = ["Procedure", "Section", "Test", "Checkpoint", "Timestamp", "% Diff", "Result", "Original", "Playback"]

        ws.append(headers)

        bold = Font(bold=True)

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)

            cell.font = bold

            cell.alignment = Alignment(horizontal="center")

        ws.freeze_panes = "A2"

        self._apply_column_widths(ws)

    def _apply_column_widths(self, ws) -> None:
        widths = [16, 14, 36, 14, 20, 10, 12, 48, 48]

        for idx, width in enumerate(widths, start=1):
            column_letter = chr(ord("A") + idx - 1)

            ws.column_dimensions[column_letter].width = width

    def _init_primary_bounds(self) -> Tuple[int, int, int, int]:
        try:
            sw, sh = pyautogui.size()

            return (0, 0, int(sw), int(sh))

        except Exception:
            return (0, 0, 1920, 1080)

    def _in_primary_monitor(self, x: Optional[int], y: Optional[int]) -> bool:
        if x is None or y is None:
            return True

        left, top, right, bottom = self._primary_bounds

        try:
            xi = int(x)

            yi = int(y)

        except Exception:
            return False

        if left <= xi < right and top <= yi < bottom:
            return True

        # Screen size might have changed; refresh once.

        self._primary_bounds = self._init_primary_bounds()

        left, top, right, bottom = self._primary_bounds

        return left <= xi < right and top <= yi < bottom

    def _select_script_path(self, script_name: str) -> Path:
        base = self.config.scripts_dir / f"{script_name}.json"
        return base

    def _play_explorer_action(self, action_dict: Dict[str, Any]) -> None:
        if ExplorerController is None:
            logger.info("Explorer automation unavailable; skipping %s", action_dict.get("action_type"))

            return

        controller = self._get_explorer_controller()

        if controller is None:
            logger.info("Explorer controller could not be initialised; skipping %s", action_dict.get("action_type"))

            return

        try:
            explorer_action = self._to_action_dataclass(action_dict)

            controller.handle(explorer_action)

        except Exception as exc:
            logger.warning("Explorer action skipped (%s): %s", action_dict.get("action_type"), exc)

    def _get_explorer_controller(self):
        if ExplorerController is None:
            return None

        if self._explorer_controller is None:
            base_path = getattr(self.config, "scripts_dir", None)

            self._explorer_controller = ExplorerController(base_path=base_path)

        return self._explorer_controller

    def _to_action_dataclass(self, payload: Dict[str, Any]) -> Action:
        field_names = getattr(Action, "__dataclass_fields__", {}).keys()

        filtered = {name: payload.get(name) for name in field_names}

        if not filtered.get("action_type"):
            filtered["action_type"] = payload.get("action_type", "explorer.unknown")

        return Action(**filtered)

    def _resolve_element_by_auto_id(
        self,
        auto_id: str,
        ctrl_type: Optional[str] = None,
        property_filter: Optional[Tuple[str, Any]] = None,
        skip_semantic: bool = False,
    ) -> Optional[Any]:
        candidates: List[Any] = []
        if not skip_semantic:
            session = self._semantic_session()
            if session is not None:
                try:
                    element = session.resolve_control(automation_id=str(auto_id), control_type=ctrl_type)
                    if element is not None:
                        candidates.append(element)
                        if self._match_property(element, property_filter):
                            return element
                except Exception as exc:
                    logger.debug("Semantic resolve failed: %s", exc)
        regex = self._normalized_title_regex()
        if Desktop is None:
            return self._first_matching_candidate(candidates, property_filter)
        self._ensure_app_window()
        try:
            desk = Desktop(backend="uia")
        except Exception as exc:
            logger.debug("Desktop backend unavailable for UIA lookup: %s", exc)
            return self._first_matching_candidate(candidates, property_filter)
        query: Dict[str, Any] = {"auto_id": auto_id}
        if ctrl_type:
            query["control_type"] = ctrl_type
        window_candidates: List[Any] = []
        if regex:
            try:
                appwin = desk.window(title_re=regex)
                try:
                    window_candidates = appwin.descendants(**query)
                except Exception:
                    window_candidates = []
                try:
                    first = appwin.child_window(**query).wrapper_object()
                    if first not in window_candidates:
                        window_candidates.append(first)
                except Exception:
                    pass
            except Exception as exc:
                self._log_uia_hint(exc)
        for candidate in window_candidates:
            candidates.append(candidate)
            if self._match_property(candidate, property_filter):
                return candidate
        try:
            fallback = desk.window(**query).wrapper_object()
            candidates.append(fallback)
            if self._match_property(fallback, property_filter):
                return fallback
        except Exception:
            try:
                fallback = desk.child_window(**query).wrapper_object()
                candidates.append(fallback)
                if self._match_property(fallback, property_filter):
                    return fallback
            except Exception as inner_exc:
                self._log_uia_hint(inner_exc)
        return self._first_matching_candidate(candidates, property_filter)

    def _first_matching_candidate(
        self,
        candidates: List[Any],
        property_filter: Optional[Tuple[str, Any]],
    ) -> Optional[Any]:
        if not candidates:
            return None
        if property_filter and property_filter[0]:
            for cand in candidates:
                if self._match_property(cand, property_filter):
                    return cand
        return candidates[0]

    def _match_property(self, element: Any, property_filter: Optional[Tuple[str, Any]]) -> bool:
        if not property_filter:
            return True
        prop_name, expected = property_filter
        if not prop_name:
            return True
        actual = self._read_element_property(element, prop_name)
        passed, _ = self._compare_property(actual, expected, "equals")
        return passed

    def _read_element_property(self, element: Any, prop: str) -> Any:
        target = (prop or "name").strip().lower()
        try:
            if target in {"text", "name", "title", "window_text"}:
                return element.window_text()
        except Exception:
            pass
        try:
            if target in {"value", "currentvalue"} and hasattr(element, "get_value"):
                return element.get_value()
        except Exception:
            pass
        try:
            if target in {"enabled", "is_enabled"}:
                return element.is_enabled()
        except Exception:
            pass
        try:
            attr = getattr(element, target, None)
            if callable(attr):
                return attr()
            if attr is not None:
                return attr
        except Exception:
            pass
        info = getattr(element, "element_info", None)
        if info is not None:
            return getattr(info, target, None)
        return None

    def _compare_property(self, actual: Any, expected: Any, comparator: str) -> tuple[bool, str]:
        actual_str = "" if actual is None else str(actual)
        expected_str = "" if expected is None else str(expected)
        note = ""
        if comparator in {"equals", "equal", "=="}:
            passed = actual_str == expected_str
            if not passed:
                note = f"expected '{expected_str}' got '{actual_str}'"
        elif comparator in {"contains", "in"}:
            passed = expected_str in actual_str
            if not passed:
                note = f"'{expected_str}' not in '{actual_str}'"
        else:
            passed = actual_str == expected_str
            if not passed:
                note = f"unknown comparator '{comparator}'"
        return passed, note

    def _record_assert_result(
        self,
        results: List[Dict[str, Any]],
        auto_id: str,
        property_name: str,
        expected: Any,
        actual: Any,
        passed: bool,
        note: str,
        semantic_meta: Optional[Dict[str, Any]] = None,
    ) -> None:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        result = {
            "index": f"assert:{auto_id}",
            "original": str(expected),
            "test": str(actual),
            "diff_percent": 0.0,
            "status": "pass" if passed else "fail",
            "timestamp": timestamp,
            "auto_id": auto_id,
            "property": property_name,
        }
        if note:
            result["note"] = note
        if semantic_meta:
            result["semantic"] = semantic_meta
        results.append(result)
        if not passed:
            self._record_failure(result["index"])



# ui_testing/testplan.py
from __future__ import annotations

import logging
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


class TestPlanReporter:
    """Update XLSM test procedures with pass/fail outcomes."""

    def __init__(self, workbook_path: Path) -> None:
        self.workbook_path = workbook_path
        self.logger = logging.getLogger(__name__)

    def mark_section(self, script_rel: str, passed: bool) -> None:
        sheet_name = self._sheet_name_for_script(script_rel)
        if not sheet_name:
            self.logger.debug("Script %s does not map to a worksheet", script_rel)
            return

        try:
            wb = openpyxl.load_workbook(self.workbook_path, keep_vba=True)
        except Exception as exc:
            self.logger.warning("Unable to open test plan %s: %s", self.workbook_path, exc)
            return

        if sheet_name not in wb.sheetnames:
            self.logger.debug("Worksheet %s not found in %s", sheet_name, self.workbook_path.name)
            wb.close()
            return

        ws = wb[sheet_name]
        status = "P" if passed else "F"
        rows_updated = self._fill_pf_column(ws, status)
        self._update_summary_sheet(wb, sheet_name, status)

        try:
            wb.save(self.workbook_path)
            self.logger.info(
                "Updated test plan %s: sheet %s marked %s (%d rows)",
                self.workbook_path.name,
                sheet_name,
                status,
                rows_updated,
            )
        except Exception as exc:
            self.logger.warning("Failed to save test plan %s: %s", self.workbook_path, exc)
        finally:
            wb.close()

    # ------------------------------------------------------------------
    def _sheet_name_for_script(self, script_rel: str) -> Optional[str]:
        parts = Path(script_rel).parts
        if len(parts) < 2:
            return None
        procedure, section = parts[0], parts[1]
        try:
            int(procedure)
            int(section)
        except ValueError:
            return None
        return f"{procedure}.{section}"

    def _fill_pf_column(self, ws: Worksheet, status: str) -> int:
        header_row, column_idx = self._locate_pf_column(ws)
        if not column_idx:
            return 0
        updates = 0
        for row in ws.iter_rows(min_row=header_row + 1, max_col=column_idx):
            cells = list(row)
            pf_cell = cells[column_idx - 1]
            other_values = [cell.value for cell in cells[: column_idx - 1]]
            if all(self._is_blank(value) for value in other_values) and self._is_blank(pf_cell.value):
                continue
            if pf_cell.value != status:
                pf_cell.value = status
                updates += 1
        return updates

    def _update_summary_sheet(self, wb: openpyxl.Workbook, sheet_name: str, status: str) -> None:
        try:
            summary_ws = wb[wb.sheetnames[0]]
        except Exception:
            return
        pf_column = None
        testno_column = None
        for cell in summary_ws[1]:
            if isinstance(cell.value, str):
                value = cell.value.strip().lower()
                if value.startswith("test no"):
                    testno_column = cell.column
                if value.startswith("pass"):
                    pf_column = cell.column
        if not pf_column or not testno_column:
            return
        for row in summary_ws.iter_rows(min_row=2, values_only=False):
            cell_value = row[testno_column - 1].value
            if isinstance(cell_value, str) and cell_value.strip().startswith(sheet_name):
                row[pf_column - 1].value = status
                break

    def _locate_pf_column(self, ws: Worksheet) -> tuple[int, Optional[int]]:
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            for col_idx, value in enumerate(row, start=1):
                if isinstance(value, str) and value.strip().lower() in {"p/f", "pass / fail", "pass/fail"}:
                    return idx, col_idx
        return ws.min_row, None

    @staticmethod
    def _is_blank(value: object) -> bool:
        if value is None:
            return True
        return isinstance(value, str) and not value.strip()

import json
import logging
import time
import threading
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Tuple, Optional
from datetime import datetime

import pyautogui
import numpy as np
from PIL import Image, ImageChops

# NEW: guarded pywinauto import (so EXE still runs even if not installed)
try:
    from pywinauto import Desktop  # UIA selector
except Exception:
    Desktop = None  # type: ignore
# EXE-safe imports
try:
    from ui_testing.action import Action  # noqa: F401
    from ui_testing.util import dotted_code_from_test_name, ensure_png_name
except Exception:
    try:
        from .action import Action  # noqa: F401
        from .util import dotted_code_from_test_name, ensure_png_name
    except Exception:
        from action import Action  # noqa: F401
        from util import dotted_code_from_test_name, ensure_png_name

logger = logging.getLogger(__name__)

@dataclass
class PlayerConfig:
    scripts_dir: Path
    images_dir: Path
    results_dir: Path
    taskbar_crop_px: int = 60
    wait_between_actions: float = 1
    app_title_regex: Optional[str] = None  # regex to scope UIA search to app window
    diff_tolerance: float = 0.01  # 0.0 = exact; >0 allows small diffs
    diff_tolerance_percent: float = 0.01
    use_default_delay_always: bool = False

class Player:
    def __init__(self, config: PlayerConfig) -> None:
        self.config = config
        self._stop_event = threading.Event()

    def play(self, script_name: str) -> List[Dict[str, Any]]:
        """Returns per-checkpoint results and writes an Excel summary per run."""
        script_path = self.config.scripts_dir / f"{script_name}.json"
        with script_path.open("r", encoding="utf-8") as f:
            actions: List[Dict[str, Any]] = json.load(f)

        results: List[Dict[str, Any]] = []
        shot_idx = 0
        base_code = dotted_code_from_test_name(Path(script_name).name)
        
        for action in actions:
            a_type = action.get("action_type")
            # Pre-action delay: prefer recorded delay, else default pacing
            if getattr(self.config, "use_default_delay_always", False):
                pre = self.config.wait_between_actions
            else:
                pre = action.get("delay", self.config.wait_between_actions)
            try:
                pre = float(pre)
            except Exception:
                pre = self.config.wait_between_actions
            pre = max(0.0, pre)
            if pre > 0:
                remaining = pre
                while remaining > 0 and not self.should_stop():
                    chunk = min(0.1, remaining)
                    time.sleep(chunk)
                    remaining -= chunk

            if self.should_stop():
                break

            if a_type == "click":
                x, y = int(action["x"]), int(action["y"])
                auto_id: Optional[str]   = action.get("auto_id")
                ctrl_type: Optional[str] = action.get("control_type")

                # Try UIA first if we have an AutomationID and pywinauto is available
                if auto_id and Desktop is not None:
                    try:
                        desk = Desktop(backend="uia")
                        target = None

                        # If a window title regex is configured, scope search within that app
                        if self.config.app_title_regex:
                            try:
                                appwin = desk.window(title_re=self.config.app_title_regex)
                                query: Dict[str, Any] = {"auto_id": auto_id}
                                if ctrl_type:
                                    query["control_type"] = ctrl_type
                                target = appwin.child_window(**query).wrapper_object()
                            except Exception:
                                target = None

                        # If scoping failed or not configured, try a global UIA search
                        if target is None:
                            query2: Dict[str, Any] = {"auto_id": auto_id}
                            if ctrl_type:
                                query2["control_type"] = ctrl_type
                            target = desk.window(**query2).wrapper_object()

                        logger.info(f"Playback(UIA): click auto_id='{auto_id}'"
                                    f"{' ctrl='+ctrl_type if ctrl_type else ''}")
                        target.click_input()
                        # Optionally wait a hair to let UI settle
                        time.sleep(0.05)
                        continue  # UIA path handled; go next action
                    except Exception as e:
                        logger.warning(
                            f"UIA click failed for auto_id='{auto_id}' (fallback to coords): {e}"
                        )

                # Fallback to screen coordinates
                logger.info(f"Playback: click at ({x}, {y})")
                pyautogui.click(x, y)

            elif a_type == "type":
                text = action.get("text", "")
                safe_preview = text.replace("\n", "<ENTER>")
                logger.info(f"Playback: type '{safe_preview}'")
                pyautogui.typewrite(text, interval=0.02)

            elif a_type == "screenshot":
                # Give the UI a brief moment to settle before capturing
                time.sleep(max(0.1, self.config.wait_between_actions))

                logger.info(f"Playback: screenshot #{shot_idx}")
                test_img = self._capture_screenshot_primary()
                img_dir = self.config.images_dir / script_name
                img_dir.mkdir(parents=True, exist_ok=True)

                # Deterministic names: 0_000T.png for playback, compare with 0_000O.png
                test_name = ensure_png_name(0, shot_idx, "T")
                test_path = img_dir / test_name
                test_img.save(test_path)

                orig_name = ensure_png_name(0, shot_idx, "O")
                orig_path = img_dir / orig_name

                passed, diff_pct = self._compare_and_highlight(orig_path, test_path)
                result = {
                    "index": shot_idx,
                    "original": str(orig_path),
                    "test": str(test_path),
                    "diff_percent": round(float(diff_pct), 3),
                    "status": "pass" if passed else "fail",
                }

                logger.info(f"Result: screenshot #{shot_idx} -> {'PASS' if passed else 'FAIL'}")
                results.append(result)
                shot_idx += 1

        # Write Excel after this test
        self._write_excel_results(script_name, results)
        return results

    def _capture_screenshot_primary(self) -> Image.Image:
        prev_failsafe = pyautogui.FAILSAFE
        pyautogui.FAILSAFE = False
        try:
            sw, sh = pyautogui.size()
            pyautogui.moveTo(sw - 5, sh - 5, duration=0)
            time.sleep(0.05)
            shot = pyautogui.screenshot()
        finally:
            pyautogui.FAILSAFE = prev_failsafe

        if self.config.taskbar_crop_px > 0:
            w, h = shot.size
            shot = shot.crop((0, 0, w, h - self.config.taskbar_crop_px))
        return shot

    def _compare_exact(self, orig_path: Path, test_path: Path) -> bool:
        """Backward-compat entry; now delegates to tolerance-based compare."""
        is_pass, _pct = self._compare_and_highlight(orig_path, test_path)
        return is_pass

    def _bounding_boxes_from_mask(self, mask: np.ndarray, cell: int = 12,
                                min_area: int = 60, pad: int = 3) -> list[tuple[int,int,int,int]]:
        """
        Cluster a boolean mask into disjoint components using a coarse grid,
        return a list of pixel-space bounding boxes (x0, y0, x1, y1).
        - cell: size of coarse grid cell in pixels (larger = faster, fewer boxes)
        - min_area: drop tiny boxes (in pixels, after final bbox)
        - pad: expand each bbox by this many pixels on each side (clamped)
        """
        h, w = mask.shape
        ys, xs = np.where(mask)
        if len(xs) == 0:
            return []

        # Map pixels to coarse cells
        gx = xs // cell
        gy = ys // cell
        cells = np.stack([gy, gx], axis=1)
        # Unique cells containing diffs
        uniq = np.unique(cells, axis=0)

        # Build adjacency on grid (8-neighborhood)
        from collections import defaultdict, deque
        cell_set = {tuple(c) for c in uniq}
        visited = set()
        boxes = []

        def cell_to_bbox(cy: int, cx: int):
            # Coarse cell bounds in pixel space
            y0 = cy * cell
            x0 = cx * cell
            y1 = min((cy + 1) * cell - 1, h - 1)
            x1 = min((cx + 1) * cell - 1, w - 1)
            return y0, x0, y1, x1

        # Precompute a fine mask to refine bounds inside coarse unions
        fine_mask = mask

        for cy, cx in uniq:
            if (cy, cx) in visited:
                continue
            # BFS over coarse neighbors
            Q = deque([(cy, cx)])
            visited.add((cy, cx))
            comp_cells = [(cy, cx)]
            while Q:
                y, x = Q.popleft()
                for dy in (-1, 0, 1):
                    for dx in (-1, 0, 1):
                        if dy == 0 and dx == 0:
                            continue
                        ny, nx = y + dy, x + dx
                        if (ny, nx) in cell_set and (ny, nx) not in visited:
                            visited.add((ny, nx))
                            Q.append((ny, nx))
                            comp_cells.append((ny, nx))

            # Convert component coarse cells to a union bbox (coarse)
            y0 = min(cy for cy, cx in comp_cells) * cell
            x0 = min(cx for cy, cx in comp_cells) * cell
            y1 = min((max(cy for cy, cx in comp_cells) + 1) * cell - 1, h - 1)
            x1 = min((max(cx for cy, cx in comp_cells) + 1) * cell - 1, w - 1)

            # Refine bbox using fine mask in that region (shrinks to true extents)
            sub = fine_mask[y0:y1+1, x0:x1+1]
            if sub.any():
                sy, sx = np.where(sub)
                ry0 = y0 + sy.min()
                ry1 = y0 + sy.max()
                rx0 = x0 + sx.min()
                rx1 = x0 + sx.max()
            else:
                ry0, ry1, rx0, rx1 = y0, y1, x0, x1

            # Pad and clamp
            ry0 = max(0, ry0 - pad); rx0 = max(0, rx0 - pad)
            ry1 = min(h - 1, ry1 + pad); rx1 = min(w - 1, rx1 + pad)

            if (ry1 - ry0 + 1) * (rx1 - rx0 + 1) >= min_area:
                boxes.append((rx0, ry0, rx1, ry1))

        return boxes

    def _compare_and_highlight(self, orig_path: Path, test_path: Path) -> tuple[bool, float]:
        """Return (is_pass, diff_percent). PASS ⇔ diff_percent <= diff_tolerance_percent.

        Also saves:
        - D image (…D.png): black & white absolute difference (all channels)
        - H image (…H.png): semi-transparent red overlay + red bounding boxes (one per region)
        """
        try:
            if not orig_path.exists() or not test_path.exists():
                logger.error(f"Missing image(s): {orig_path} | {test_path}")
                return False, 100.0

            o = Image.open(orig_path).convert("RGBA")
            t = Image.open(test_path).convert("RGBA")
            if o.size != t.size:
                t = t.resize(o.size, Image.LANCZOS)

            a = np.asarray(o, dtype=np.int16)
            b = np.asarray(t, dtype=np.int16)

            absdiff = np.abs(a - b)
            diff_mask = np.any(absdiff > 0, axis=2)      # H×W bool

            total = diff_mask.size
            num_diff = int(diff_mask.sum())
            diff_percent = (num_diff / total) * 100.0

            # ----- Save D (black & white difference) -----
            perpix = absdiff[..., :3].max(axis=2)  # ignore alpha for D
            if perpix.max() > 0:
                perpix = (perpix.astype(np.float32) / perpix.max()) * 255.0
            d_img = Image.fromarray(perpix.astype(np.uint8), mode="L")

            stem = test_path.stem
            d_name = (stem[:-1] + "D") if stem and stem[-1] in ("T", "O") else (stem + "_D")
            d_path = test_path.with_name(d_name + ".png")
            try:
                d_img.save(d_path)
            except Exception:
                pass

            # ----- Save H (semi-transparent red + MULTI boxes) -----
            if num_diff > 0:
                overlay = np.zeros_like(a)
                overlay[..., 0] = 255  # red
                overlay[..., 3] = np.where(diff_mask, 96, 0)
                hi = Image.alpha_composite(o, Image.fromarray(overlay.astype(np.uint8)))

                # draw multiple rectangles
                boxes = self._bounding_boxes_from_mask(diff_mask, cell=12, min_area=60, pad=3)
                arr = np.array(hi)
                for (x0, y0, x1, y1) in boxes:
                    # 3px outline
                    arr[y0:y0+3, x0:x1+1] = [255, 0, 0, 255]
                    arr[y1-2:y1+1, x0:x1+1] = [255, 0, 0, 255]
                    arr[y0:y1+1, x0:x0+3] = [255, 0, 0, 255]
                    arr[y0:y1+1, x1-2:x1+1] = [255, 0, 0, 255]
                hi = Image.fromarray(arr, mode="RGBA")

                h_name = (stem[:-1] + "H") if stem and stem[-1] in ("T", "O") else (stem + "_H")
                try:
                    hi.save(test_path.with_name(h_name + ".png"))
                except Exception:
                    pass

            # ----- PASS/FAIL based on tolerance -----
            tol = float(getattr(self.config, "diff_tolerance_percent",
                                getattr(self.config, "diff_tolerance", 0.0)))
            is_pass = (diff_percent <= tol)
            return is_pass, diff_percent

        except Exception as e:
            logger.exception(f"Compare failed for {orig_path} vs {test_path}: {e}")
            return False, 100.0

    def _make_highlight_image(self, base_img: Image.Image, mask: np.ndarray) -> Image.Image:
        """Return base_img with semi-transparent red overlay where mask is True, and a union bbox outline."""
        h, w = mask.shape
        # Create an RGBA from base
        comp = base_img.convert("RGBA")
        overlay = Image.new("RGBA", (w, h), (255, 0, 0, 0))
        # Semi-transparent red where mask True
        alpha = 96  # tweak intensity (0..255)
        r = np.zeros((h, w, 4), dtype=np.uint8)
        r[..., 0] = 255  # R
        r[..., 3] = 0
        r[mask, 3] = alpha
        overlay = Image.fromarray(r, mode="RGBA")
        comp = Image.alpha_composite(comp, overlay)

        # Draw a union bounding box (single outline) for quick spotting
        if mask.any():
            ys, xs = np.where(mask)
            y0, y1 = ys.min(), ys.max()
            x0, x1 = xs.min(), xs.max()
            # Draw rectangle  (simple 3px outline)
            draw = Image.fromarray(np.array(comp))
            arr = np.array(draw)
            # top & bottom
            arr[y0:y0+3, x0:x1+1] = [255, 0, 0, 255]
            arr[y1-2:y1+1, x0:x1+1] = [255, 0, 0, 255]
            # left & right
            arr[y0:y1+1, x0:x0+3] = [255, 0, 0, 255]
            arr[y0:y1+1, x1-2:x1+1] = [255, 0, 0, 255]
            comp = Image.fromarray(arr, mode="RGBA")

        return comp.convert("RGB")

    def _split_hierarchy(self, script_name: str) -> Tuple[str, str, str]:
        p = Path(script_name)
        test = p.name
        sec = p.parent.name if p.parent.name else ""
        proc = p.parent.parent.name if p.parent.parent and p.parent.parent.name else ""
        return proc, sec, test

    def request_stop(self, clear_only: bool=False):
        """If clear_only=True, clears previous stop; else sets stop."""
        if clear_only:
            self._stop_event.clear()
        else:
            self._stop_event.set()

    def should_stop(self) -> bool:
        return self._stop_event.is_set()

    def _write_excel_results(self, script_name: str, results: List[Dict[str, Any]]) -> None:
        # Lazy import so users without Excel don't break other features
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except Exception as e:
            logger.warning(f"Excel export skipped (openpyxl not available): {e}")
            return

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = self.config.results_dir / f"results_{ts}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Results"

        headers = ["Procedure", "Section", "Test", "Checkpoint", "Original", "Test", "% Diff", "Result"]

        ws.append(headers)
        bold = Font(bold=True)
        for col in range(1, len(headers)+1):
            ws.cell(row=1, column=col).font = bold
            ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

        proc, sec, test = self._split_hierarchy(script_name)

        for r in results:
            checkpoint = f"{r.get('index',0)}"
            row = [
                proc, sec, test,
                checkpoint,
                r.get("original",""),
                r.get("test",""),
                round(float(r.get("diff_percent", 0.0)), 3),
                "PASS" if r.get("status") == "pass" else "FAIL",
            ]
            ws.append(row)
            # Color result cell
            last_row = ws.max_row
            cell = ws.cell(row=last_row, column=8)
            if row[-1] == "PASS":
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # width tuning
        widths = [14, 10, 34, 12, 46, 46, 10, 10]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(ord('A') + i - 1)].width = w

        try:
            wb.save(out_path)
            logger.info(f"Excel results saved: {out_path}")
        except Exception as e:
            logger.warning(f"Failed to save Excel results: {e}")
